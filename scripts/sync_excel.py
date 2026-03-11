#!/usr/bin/env python3
import argparse
import datetime as dt
import hashlib
import json
import os
import subprocess
import sys
from pathlib import Path


def stable_hash(obj) -> str:
    data = json.dumps(obj, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(data.encode("utf-8")).hexdigest()


def load_sheet(excel_path: Path, sheet_name: str):
    try:
        from openpyxl import load_workbook
    except Exception:
        print("Missing dependency: openpyxl. Install with: python3 -m pip install openpyxl", file=sys.stderr)
        raise

    wb = load_workbook(excel_path, data_only=True, read_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return [], []

    headers = []
    for i, h in enumerate(raw_headers, start=1):
        name = str(h).strip() if h is not None else ""
        headers.append(name or f"Column_{i}")

    records = []
    for row in rows_iter:
        if row is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        obj = {}
        for idx, header in enumerate(headers):
            val = row[idx] if idx < len(row) else None
            if isinstance(val, (dt.datetime, dt.date, dt.time)):
                val = val.isoformat()
            obj[header] = val
        records.append(obj)

    return headers, records


def git_commit(repo_root: Path, message: str):
    subprocess.run(["git", "add", "data/data.json", ".state/sync_state.json"], cwd=repo_root, check=True)

    diff = subprocess.run(
        ["git", "diff", "--cached", "--quiet"],
        cwd=repo_root,
        check=False,
    )
    if diff.returncode == 0:
        print("No staged changes to commit.")
        return False

    subprocess.run(["git", "commit", "-m", message], cwd=repo_root, check=True)
    subprocess.run(["git", "push"], cwd=repo_root, check=True)
    return True


def main():
    parser = argparse.ArgumentParser(description="Sync Excel sheet to dashboard JSON")
    parser.add_argument("--excel", required=True, help="Path to .xlsx/.xlsm file")
    parser.add_argument("--sheet", required=True, help="Sheet name")
    parser.add_argument("--repo", default=".", help="Dashboard repo root")
    parser.add_argument("--commit", action="store_true", help="Commit and push when changed")
    args = parser.parse_args()

    repo_root = Path(args.repo).resolve()
    data_path = repo_root / "data" / "data.json"
    state_path = repo_root / ".state" / "sync_state.json"
    data_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.parent.mkdir(parents=True, exist_ok=True)

    excel_path = Path(os.path.expanduser(args.excel)).resolve()
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    headers, rows = load_sheet(excel_path, args.sheet)

    # Read current model score from Portfolios tab J17
    model_score = None
    try:
        wb_score = load_workbook(excel_path, data_only=True, read_only=True, keep_vba=True)
        model_score = wb_score["Portfolios"].cell(row=17, column=10).value
    except Exception as e:
        print(f"Warning: could not read model score: {e}", file=sys.stderr)

    payload_core = {
        "sheet": args.sheet,
        "source": str(excel_path),
        "headers": headers,
        "rows": rows,
        "modelScore": model_score,
    }
    content_hash = stable_hash(payload_core)

    prev_hash = None
    if state_path.exists():
        try:
            prev_hash = json.loads(state_path.read_text(encoding="utf-8")).get("contentHash")
        except Exception:
            prev_hash = None

    if content_hash == prev_hash:
        print("No data change detected.")
        return

    payload = {
        **payload_core,
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "rowCount": len(rows),
    }

    data_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    # Also write data.js so index.html works when opened directly (file://)
    js_path = repo_root / "data" / "data.js"
    js_path.write_text(
        "window.DASHBOARD_DATA = " + json.dumps(payload, indent=2, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )

    state_path.write_text(
        json.dumps({"contentHash": content_hash, "updatedAt": payload["generatedAt"]}, indent=2),
        encoding="utf-8",
    )
    print(f"Updated {data_path} and data/data.js with {len(rows)} rows.")

    if args.commit:
        try:
            committed = git_commit(repo_root, f"Update dashboard data ({args.sheet})")
            if committed:
                print("Committed and pushed changes.")
        except subprocess.CalledProcessError as e:
            print(f"Git command failed: {e}", file=sys.stderr)
            sys.exit(e.returncode)


if __name__ == "__main__":
    main()
